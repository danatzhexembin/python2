import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numpy as np

file_name_1 = "Лист1.xlsx"
file_name_2 = "Лист2.xlsx"
file_name_3 = "Лист3.xlsx"

# загружаем в память уже существующий файл на диске
workbook_1 = openpyxl.load_workbook(file_name_1)
workbook_2 = openpyxl.load_workbook(file_name_2)
workbook_3 = openpyxl.load_workbook(file_name_3)

# берём активную страницу из рабочей книги
worksheet_1 = workbook_1.active
worksheet_2 = workbook_2.active
worksheet_3 = workbook_3.active


# для первой матрицы
# берём последнюю строку в excel-файле
max_row_1 = worksheet_1.max_row + 1
# берём последнюю колонку в excel-файле
max_column_1 = worksheet_1.max_column + 1

# количество столбцов
m = max_column_1 - 1
# количество строк
n = max_row_1 - 1
# первые [] - строка, вторые [] - столбец
matrix_1 = [[0] * m for i in range(n)]

for i in range(1, max_row_1):
    for j in range(1, max_column_1):

        value = worksheet_1.cell(row=i, column=j).value
        matrix_1[i - 1][j - 1] = value
        # print(f"{i} {j}\n{value}\n")
print("MATRIX 1:")
print(matrix_1)


# для второй матрицы
# берём последнюю строку в excel-файле
max_row_2 = worksheet_2.max_row + 1
# берём последнюю колонку в excel-файле
max_column_2 = worksheet_2.max_column + 1

# количество столбцов
m = max_column_2 - 1
# количество строк
n = max_row_2 - 1
# первые [] - строка, вторые [] - столбец
matrix_2 = [[0] * m for i in range(n)]

for i in range(1, max_row_2):
    for j in range(1, max_column_2):

        value = worksheet_2.cell(row=i, column=j).value
        matrix_2[i - 1][j - 1] = value
        # print(f"{i} {j}\n{value}\n")
print("MATRIX 2:")
print(matrix_2)


# для третьей матрицы
# берём последнюю строку в excel-файле
max_row_3 = worksheet_3.max_row + 1
# берём последнюю колонку в excel-файле
max_column_3 = worksheet_3.max_column + 1

# количество столбцов
m = max_column_3 - 1
# количество строк
n = max_row_3 - 1
# первые [] - строка, вторые [] - столбец
matrix_3 = [[0] * m for i in range(n)]

for i in range(1, max_row_3):
    for j in range(1, max_column_3):

        value = worksheet_3.cell(row=i, column=j).value
        matrix_3[i - 1][j - 1] = value
        # print(f"{i} {j}\n{value}\n")
print("MATRIX 3:")
print(matrix_3)


# Нужно создавать новый Workbook для записи новых данных
workbook_new = Workbook()

worksheet_new = workbook_new.active

# запись в файл первой матрицы
for i in range(1, max_row_1):
    for j in range(1, max_column_1):
        row = i
        # value = worksheet.cell(row=i, column=j).value
        value1 = matrix_1[i - 1][j - 1]
        col = get_column_letter(j)
        # записываем значение в выбранную ячейку
        worksheet_new[f'{col}{row}'] = str(value1)

# запись в файл второй матрицы
for i in range(max_row_1 + 1, max_row_2):
    for j in range(max_column_1 + 1, max_column_2):
        row = i
        # value = worksheet.cell(row=i, column=j).value
        value1 = matrix_2[i - 1][j - 1]
        col = get_column_letter(j)
        # записываем значение в выбранную ячейку
        worksheet_new[f'{col}{row}'] = str(value1)

workbook_new.save("result.xlsx")