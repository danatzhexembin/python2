import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numpy as np

file_name = "task.xlsx"

# загружаем в память уже существующий файл на диске
workbook = openpyxl.load_workbook(file_name)

# берём активную страницу из рабочей книги
worksheet = workbook.active

# берём последнюю строку в excel-файле
max_row = worksheet.max_row + 1
# берём последнюю колонку в excel-файле
max_column = worksheet.max_column + 1

# количество столбцов
m = max_column - 1
# количество строк
n = max_row - 1
# первые [] - строка, вторые [] - столбец
matrix = [[0] * m for i in range(n)]
# matrix[1][2] = 3
# print(matrix)

for i in range(1, max_row):
    for j in range(1, max_column):

        value = worksheet.cell(row=i, column=j).value
        matrix[i - 1][j - 1] = value
        # print(f"{i} {j}\n{value}\n")
print(matrix)

# Нужно создавать новый Workbook для записи новых данных
workbook_new = Workbook()

worksheet_new = workbook_new.active

for i in range(1, max_row):
    for j in range(1, max_column):
        row = j
        # value = worksheet.cell(row=i, column=j).value
        value1 = matrix[i - 1][j - 1]
        col = get_column_letter(i)
        # записываем значение в выбранную ячейку
        worksheet_new[f'{col}{row}'] = str(value1)

workbook_new.save("result.xlsx")
