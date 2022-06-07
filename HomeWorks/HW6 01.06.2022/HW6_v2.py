import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def function_len_array(array):  # ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']
    return len(array)


file_name_1 = "Лист1.xlsx"
file_name_2 = "Лист2.xlsx"
file_name_3 = "Лист3.xlsx"

# загружаем в память уже существующий файл на диске
workbook_1 = openpyxl.load_workbook(file_name_1)
workbook_2 = openpyxl.load_workbook(file_name_2)
workbook_3 = openpyxl.load_workbook(file_name_3)

# берём активную страницу из рабочей книги
worksheet_1 = workbook_1.active
worksheet_2 = workbook_2.active
worksheet_3 = workbook_3.active


# для первой матрицы
max_row_1 = worksheet_1.max_row + 1
matrix_1 = [0 for i in range(max_row_1 - 1)]

for i in range(1, max_row_1):
    value = worksheet_1.cell(row=i, column=1).value
    matrix_1[i - 1] = value
    # print(f"{i} {j}\n{value}\n")
# print("MATRIX 1:")
# print(matrix_1)


# для второй матрицы
# берём последнюю строку в excel-файле
max_row_2 = worksheet_2.max_row + 1

# первые [] - строка, вторые [] - столбец
matrix_2 = [0 for i in range(max_row_2 - 1)]

for i in range(1, max_row_2):
    value = worksheet_2.cell(row=i, column=2).value
    matrix_2[i - 1] = value
    # print(f"{i} {j}\n{value}\n")
# print("MATRIX 2:")
# print(matrix_2)


# для третьей матрицы
# берём последнюю строку в excel-файле
max_row_3 = worksheet_3.max_row + 1

# первые [] - строка, вторые [] - столбец
matrix_3 = [0 for i in range(max_row_3 - 1)]

for i in range(1, max_row_3):
    value = worksheet_3.cell(row=i, column=3).value
    matrix_3[i - 1] = value
    # print(f"{i} {j}\n{value}\n")
# print("MATRIX 3:")
# print(matrix_3)

matrix_all = [matrix_1, matrix_2, matrix_3]

print(matrix_all)

# Нужно создавать новый Workbook для записи новых данных
workbook_new = Workbook()

sheet = workbook_new.active
workbook_new.remove(sheet)

workbook_new.create_sheet("ALL")
worksheet_new = workbook_new["ALL"]

for row in range(0, function_len_array(matrix_all)):  # [0, 1, 2, 3, 4, ..., 1007]
    # print(f"col_count: {len(external_array[row])}")
    for col in range(0, function_len_array(matrix_all[row])):
        # [0, 1, 2, 3, 4, 5]  # external_array[row] == ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']

        worksheet_new[f'{get_column_letter(row + 1)}{col + 1}'] = matrix_all[row][col]  # записали в виде столбца (а
        # были матрицы строки)
        # if row == 0:
        #     worksheet_new[f'{get_column_letter(row + 1)}{col + 1}'].font = Font(bold=True) #  жирный текст
        # pass
    # print(external_array[row])
    pass

workbook_new.create_sheet("Лист1")
worksheet_list1 = workbook_new["Лист1"]

for col in range(0, function_len_array(matrix_all[0])):
    worksheet_list1[f'{get_column_letter(1)}{col + 1}'] = matrix_all[0][col]

workbook_new.create_sheet("Лист2")
worksheet_list2 = workbook_new["Лист2"]

for col in range(0, function_len_array(matrix_all[1])):
    worksheet_list2[f'{get_column_letter(2)}{col + 1}'] = matrix_all[1][col]

workbook_new.create_sheet("Лист3")
worksheet_list3 = workbook_new["Лист3"]

for col in range(0, function_len_array(matrix_all[2])):
    worksheet_list3[f'{get_column_letter(3)}{col + 1}'] = matrix_all[2][col]

workbook_new.save("result.xlsx")
